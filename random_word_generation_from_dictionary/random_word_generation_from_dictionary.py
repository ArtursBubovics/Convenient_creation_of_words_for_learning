from openpyxl.styles import PatternFill
import random

def random_word_generation_from_dictionary(excel_file, df, workbook, worksheet):
        generate_word = True

        place_of_generated_words = []

        for index, row in df.iterrows():
            excel_row = index + 2  # +3 для учета заголовков и начальной строки

            # Индекс столбца в Excel для 'Is in cards' (начиная с B, которая является 2-м столбцом)
            is_in_cards_column = 2

            is_in_cards_cell = worksheet.cell(row=excel_row, column=is_in_cards_column)

            if cell_color_determination (is_in_cards_cell, 'FF0070C0'):
                #print('blue')
                print("Дальше нету полей для ввода")
                break
            else:
                #print('green')
                if cell_color_determination (is_in_cards_cell, 'FF00B050'):
                    place_of_generated_words.append((excel_row, is_in_cards_column))
                    continue
                else:
                     continue
                    
        #print(place_of_generated_words)

        while generate_word:
            if not place_of_generated_words:
                print("Нет доступных слов для генерации")
                break

            row_value, column_value = random.choice(place_of_generated_words)
            
            print(f"\n\nВот сгенерированое слово: \033[92m{worksheet.cell(row=row_value, column=column_value+2).value}\033[0m")

            user_response = input('\nДальше генерировать?\n Да это + ; Нет это - ;  Выберите действие: ')

            if(user_response == '-'):
                if(input('\nХотите сохранить изменения?\n Да это + ; Нет это - ; Выберите действие: ') == '+'):
                    worksheet.cell(row=row_value, column=column_value+1).value = worksheet.cell(row=row_value, column=column_value+1).value + 1
                    workbook.save(excel_file)
                    break
                else:
                    break
            else:
                worksheet.cell(row=row_value, column=column_value+1).value = worksheet.cell(row=row_value, column=column_value+1).value + 1


def cell_color_determination(cell, color):
    fill = cell.fill
    #print(f"Cell fill color: {fill.start_color.index}")
    return fill.start_color.index == color