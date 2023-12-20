import re;
from response_generation import response_generation
from openpyxl.styles import PatternFill

def response_func(search_word, search_word_translate, df,index,worksheet):
    
    generated_sentences = []
    last_match = None
    
    while True:

        if not generated_sentences:
            response = response_generation(1,generated_sentences, search_word, search_word_translate)
        else:
            response = response_generation(2,generated_sentences, search_word, search_word_translate)
    
        # Вывод ответа
        print()  
        print(response)
        print() 

        # Предложение для сохранения
        user_input = input('Введите "+" для сохранения, "-" для повторного запроса, "/" для остановки работы и сохранения изменений: ')

        # Паттерн для поиска текста внутри *
        pattern = r'\*([^*]*)\*'

        # Используем re.search для поиска первого соответствия
        matches = re.findall(pattern, response)

        # Получаем последнее вхождение
        if matches:
            last_match = matches[-1]
        else:
            print("Совпадений не найдено.")

        if user_input == '+':
            # Сохранение предложения в другом поле в Excel
            df.at[index, 'примеры в тексте'] = response.strip()
            
            # Сохранение обновленных данных в Excel
            try:

                # Если найдено соответствие, получаем текст внутри *
                if last_match:
                    worksheet.cell(row=index + 2, column=df.columns.get_loc('примеры в тексте') + 1, value=last_match.strip())
                    green_fill = PatternFill(start_color='00B050', end_color='00FF00', fill_type='solid')
                    worksheet.cell(row=index + 2, column=df.columns.get_loc('В картачках') + 1).fill = green_fill
                    print('\033[91mДанные успешно добавлены в буфер. Для сохранения нужно нажать на " / ", чтобы все сохранить\033[0m')
                    print()
                    print('---------------------')
                    print()
                    generated_sentences = []
                    return False
                else:
                    print("Соответствие не найдено.1")
                    return False

            except Exception as e:
                print(f"Произошла ошибка при сохранении в Excel: {e}")
        elif user_input == '-':
            if last_match:
                generated_sentences.append(last_match)

            else:
                print("Соответствие не найдено.2")

            continue    # Повторить запрос, пропустив текущую итерацию цикла

        elif user_input == '/':
            generated_sentences = []
            return True
        else:
            print("Некорректный ввод. Повторите попытку.")