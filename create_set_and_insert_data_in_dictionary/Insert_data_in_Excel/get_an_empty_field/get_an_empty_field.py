from openpyxl.styles import PatternFill
import pandas as pd

BLUE_COLOR = 'FF0070C0'

def get_an_empty_field(excel_file, df, workbook, worksheet):

    # Проход по каждой строке в Excel
    for index, row in df.iterrows():
        print()
        excel_row = index + 2  # +3 для учета заголовков и начальной строки

        # Индекс столбца в Excel для 'Is in cards' (начиная с B, которая является 2-м столбцом)
        is_in_cards_column = 2

        is_in_cards_cell = worksheet.cell(row=excel_row, column=is_in_cards_column)
        
        if cell_color_determination (is_in_cards_cell, BLUE_COLOR):
            print('blue')
            print("Дальше нету полей для ввода")
            break
        else:
            print('white')
            if not cell_has_no_fill (is_in_cards_cell):
                print('go1')
                if (not pd.isna(worksheet.cell(row=excel_row, column=3)) and 
                    not pd.isna(worksheet.cell(row=excel_row, column=4)) and 
                    not pd.isna(worksheet.cell(row=excel_row, column=5)) and 
                    not pd.isna(worksheet.cell(row=excel_row, column=6)) and 
                    not pd.isna(worksheet.cell(row=excel_row, column=7)) and 
                    not pd.isna(worksheet.cell(row=excel_row, column=8))):
                    print(worksheet.cell(row=excel_row, column=2))
                    return worksheet.cell(row=excel_row, column=2)
                else:
                    continue
            else:
                continue


def cell_color_determination(cell, color):
    fill = cell.fill
    print(f"Cell fill color: {fill.start_color.index}")
    if isinstance(fill, PatternFill):
        return fill.start_color.index == color

def cell_has_no_fill(cell):
    fill = cell.fill
    return fill is None